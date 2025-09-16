from .models import ExaminationRecord, Notification
from rest_framework.permissions import IsAuthenticated
from django.views import View
from rest_framework import generics, status, filters
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework_simplejwt.views import TokenObtainPairView, TokenRefreshView
from rest_framework_simplejwt.tokens import RefreshToken
from .models import (
    User,
    Appointment,
    ActivityLog,
    SystemSettings,
    PatientHealthRecord,
    ExaminationRecord,
    PreviousPregnancy,
    Notification,
    AppointmentReminder
)
from .serializers import (
    CustomTokenObtainPairSerializer,
    AdminCreateUserSerializer,
    UserSerializer,
    AdminUpdateUserSerializer,
    AppointmentSerializer,
    AppointmentCreateSerializer,
    ActivityLogSerializer,
    AdminProfileSerializer,
    SystemSettingsSerializer,
    ProviderAddMotherSerializer,
    ProviderUpdateMotherSerializer,
    MotherListSerializer,
    ProviderMotherDetailSerializer,
    PatientHealthRecordSerializer,
    ExaminationRecordSerializer,
    PreviousPregnancySerializer,
    NotificationSerializer,
    AppointmentReminderSerializer,
    MotherContactUpdateSerializer,
    PatientHealthRecordSerializer,
    MotherProfileSerializer,
    ChangePasswordSerializer,
    ProviderProfileSerializer,
    ProviderDashboardSerializer,
    UnifiedAppointmentSerializer,
)
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.decorators import api_view, permission_classes
from .permissions import IsAdmin, IsHealthcareProvider, IsMother
from rest_framework.generics import UpdateAPIView, ListAPIView, RetrieveUpdateAPIView
from django.contrib.auth.password_validation import validate_password
from rest_framework.serializers import Serializer, CharField, ValidationError
import io
import csv
from django.http import HttpResponse
from rest_framework.renderers import JSONRenderer
from django.db import models
from django.db.models import Q, Prefetch, Count
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter
from drf_spectacular.utils import extend_schema, extend_schema_view, OpenApiResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from io import BytesIO
from rest_framework import status as drf_status
from rest_framework.pagination import PageNumberPagination
from .filters import ActivityLogFilter
from rest_framework.exceptions import PermissionDenied, ValidationError
from django.shortcuts import get_object_or_404
from .utils import generate_custom_id, generate_password
from datetime import datetime, timedelta
from .filters import MotherFilter
from django.utils.dateparse import parse_date
from django.utils import timezone
from django.contrib.auth import get_user_model
from django.utils.timezone import now
from django.contrib.auth.hashers import check_password

User = get_user_model()


@extend_schema_view(
    post=extend_schema(
        summary="Obtain JWT Token",
        tags=["Authentication"]
    )
)
class CustomTokenObtainPairView(TokenObtainPairView):
    permission_classes = [AllowAny]
    serializer_class = CustomTokenObtainPairSerializer

    def logout_view(request):
        try:
            token = RefreshToken(request.data["refresh"])
            token.blacklist()
            return Response({"detail": "Logged out successfully"})
        except Exception:
            return Response({"error": "Invalid token"}, status=400)

    def get_me(request):
        return Response(UserSerializer(request.user).data)


@extend_schema_view(
    get=extend_schema(
        summary="Welcome message for admin",
        tags=["Admin"],
        responses=OpenApiResponse(description="Welcome message for admin")
    )
)
class AdminOnlyView(APIView):
    permission_classes = [IsAdmin]

    def get(self, request):
        return Response({"message": "Welcome, Admin!"})


class AdminProfileView(RetrieveUpdateAPIView):
    permission_classes = [IsAuthenticated, IsAdmin]
    serializer_class = AdminProfileSerializer

    def get_object(self):
        return self.request.user


class AdminProfileUpdateView(UpdateAPIView):
    serializer_class = AdminUpdateUserSerializer
    permission_classes = [IsAuthenticated, IsAdmin]

    def get_object(self):
        return self.request.user

    def get(self, request, *args, **kwargs):
        serializer = self.get_serializer(self.get_object())
        return Response(serializer.data, status=status.HTTP_200_OK)


@extend_schema_view(
    get=extend_schema(
        summary="Provider-only welcome message",
        tags=["Provider"]
    )
)
class ProviderOnlyView(APIView):
    permission_classes = [IsHealthcareProvider]

    def get(self, request):
        return Response({"message": "Welcome, Healthcare Provider!"})


@extend_schema_view(
    get=extend_schema(
        summary="Mother-only welcome message",
        tags=["Mother"]
    )
)
class MotherOnlyView(APIView):
    permission_classes = [IsMother]

    def get(self, request):
        return Response({"message": "Welcome, Mother!"})


@extend_schema_view(
    post=extend_schema(
        summary="Create user (Admin only)",
        tags=["User Management"]
    )
)
class AdminCreateUserView(generics.CreateAPIView):
    queryset = User.objects.all()
    serializer_class = AdminCreateUserSerializer
    permission_classes = [IsAuthenticated, IsAdmin]

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        # Save user instance and retrieve credentials
        user = serializer.save()
        ActivityLog.objects.create(
            actor=request.user,
            action="Created user",
            description=f"{request.user.username} created user {user.username}"
        )

        # These values were auto-generated in the serializer
        username = user.username
        first = user.first_name[:3]
        last = user.last_name[:3]
        password = (first + last + '123').lower()

        return Response({
            'message': 'User created successfully.',
            'username': username,
            'custom_id': user.custom_id,
            'password': password
        }, status=status.HTTP_201_CREATED)


@extend_schema_view(
    put=extend_schema(
        summary="Change password",
        tags=["User Management"]
    )
)
class ChangePasswordView(UpdateAPIView):
    serializer_class = ChangePasswordSerializer
    permission_classes = [IsAuthenticated]

    def get_object(self):
        return self.request.user

    def update(self, request, *args, **kwargs):
        user = self.get_object()
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        if not user.check_password(serializer.validated_data['old_password']):
            raise ValidationError({'old_password': 'Incorrect old password.'})

        user.set_password(serializer.validated_data['new_password'])
        user.save()

        return Response({'detail': 'Password updated successfully.'}, status=status.HTTP_200_OK)


@extend_schema_view(
    get=extend_schema(
        summary="List users with filters & export",
        tags=["User Management"]
    )
)
class UserListView(generics.ListAPIView):
    queryset = User.objects.all().order_by('-id')
    serializer_class = UserSerializer
    permission_classes = [IsAuthenticated, IsAdmin]
    renderer_classes = [JSONRenderer]  # default renderer

    def get_queryset(self):
        queryset = super().get_queryset()
        role = self.request.query_params.get('role')
        search = self.request.query_params.get('search')

        if role:
            queryset = queryset.filter(role__iexact=role)

        if search:
            queryset = queryset.filter(
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(username__icontains=search) |
                Q(email__icontains=search) |
                Q(phone_number__icontains=search) |
                Q(custom_id__icontains=search)
            )

        return queryset

    def list(self, request, *args, **kwargs):
        export_format = request.query_params.get('format')

        # CSV/Excel/PDF export
        if export_format in ['csv', 'excel', 'pdf']:
            queryset = self.filter_queryset(self.get_queryset())
            serializer = self.get_serializer(queryset, many=True)

            if export_format == 'csv':
                return self._export_csv(serializer.data)
            elif export_format == 'excel':
                return self._export_excel(serializer.data)
            elif export_format == 'pdf':
                return self._export_pdf(serializer.data)

            # You can implement excel/pdf here if needed
            return Response({'error': 'Only CSV export is implemented.'}, status=400)

        return super().list(request, *args, **kwargs)

    def _export_csv(self, data):
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        if data:
            writer.writerow(data[0].keys())  # headers
            for row in data:
                writer.writerow(row.values())

        buffer.seek(0)
        response = HttpResponse(buffer, content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename=users.csv'
        return response


@extend_schema_view(
    delete=extend_schema(
        summary="Delete user by ID",
        tags=["User Management"]
    )
)
class DeleteUserView(APIView):
    permission_classes = [IsAuthenticated, IsAdmin]

    def delete(self, request, user_id):
        try:
            user = User.objects.get(id=user_id)
            user.delete()
            ActivityLog.objects.create(
                actor=request.user,
                action="Deleted user",
                description=f"{request.user.username} deleted user {user.username}"
            )

            return Response({"detail": "User deleted successfully."}, status=status.HTTP_204_NO_CONTENT)
        except User.DoesNotExist:
            return Response({"error": "User not found."}, status=status.HTTP_404_NOT_FOUND)


@extend_schema_view(
    put=extend_schema(
        summary="Update user info (Admin only)",
        tags=["User Management"]
    )
)
class AdminUpdateUserView(generics.UpdateAPIView):
    queryset = User.objects.all()
    serializer_class = (
        AdminUpdateUserSerializer,
        # UserSerializer
    )
    permission_classes = [IsAuthenticated, IsAdmin]
    lookup_url_kwarg = "user_id"

    def get_object(self):
        user_id = self.kwargs.get("user_id")
        try:
            return User.objects.get(id=user_id)
        except User.DoesNotExist:
            from rest_framework.exceptions import NotFound
            raise NotFound("User not found.", status=status.HTTP_404_NOT_FOUND)

    def update(self, request, *args, **kwargs):
        user = self.get_object()  # The user being updated
        response = super().update(request, *args, **kwargs)

        # Log the update action
        ActivityLog.objects.create(
            actor=request.user,  # the admin doing the update
            action="Updated user",
            description=f"{request.user.username} updated user {user.username}"
        )

        return response


@extend_schema_view(
    post=extend_schema(
        summary="Create appointment",
        tags=["Appointments"]
    )
)
class CreateAppointmentView(generics.CreateAPIView):
    queryset = Appointment.objects.all()
    serializer_class = AppointmentCreateSerializer
    permission_classes = [IsAuthenticated, IsAdmin | IsHealthcareProvider]

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        if not serializer.is_valid():
            print("❌ Appointment creation failed:",
                  serializer.errors)  # Add this line
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        self.perform_create(serializer)
        ActivityLog.objects.create(
            actor=request.user,
            action="Created appointment",
            description=f"{request.user.username} created appointment for {serializer.data.get('patient_name')}"
        )

        return Response(serializer.data, status=status.HTTP_201_CREATED)


@extend_schema_view(
    get=extend_schema(
        summary="List 10 recent appointments",
        tags=["Appointments"]
    )
)
class RecentAppointmentsView(generics.ListAPIView):
    serializer_class = AppointmentSerializer
    permission_classes = [IsAuthenticated, IsAdmin]

    def get_queryset(self):
        recent = Appointment.objects.order_by('-appointment_date')[:10]
        return recent


@extend_schema_view(
    get=extend_schema(
        summary="List all appointments with filters & export",
        tags=["Appointments"]
    )
)
class AppointmentListView(generics.ListAPIView):
    queryset = Appointment.objects.all().order_by('-appointment_date')
    serializer_class = AppointmentSerializer
    permission_classes = [IsAuthenticated, IsAdmin | IsHealthcareProvider]
    filter_backends = [DjangoFilterBackend, SearchFilter]
    filterset_fields = ['appointment_date']
    search_fields = ['appointment_type', 'patient__first_name',
                     'patient__last_name', 'provider__first_name', 'provider__last_name']

    def list(self, request, *args, **kwargs):
        export_format = request.query_params.get('format')

        # If exporting
        if export_format in ['csv', 'excel', 'pdf']:
            queryset = self.filter_queryset(self.get_queryset())
            serializer = self.get_serializer(queryset, many=True)

            if export_format == 'csv':
                return self._export_csv(serializer.data)
            elif export_format == 'excel':
                return self._export_excel(serializer.data)
            elif export_format == 'pdf':
                return self._export_pdf(serializer.data)

            return Response({'error': 'Unsupported export format.'}, status=400)

        return super().list(request, *args, **kwargs)

    def _export_csv(self, data):
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        if data:
            writer.writerow(data[0].keys())  # header
            for row in data:
                writer.writerow(row.values())

        buffer.seek(0)
        response = HttpResponse(buffer, content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename=appointments.csv'
        return response

    def _export_excel(self, data):
        wb = Workbook()
        ws = wb.active
        ws.append(list(data[0].keys()) if data else [])  # headers
        for row in data:
            ws.append(list(row.values()))

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=appointments.xlsx'
        return response

    def _export_pdf(self, data):
        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=letter)
        width, height = letter
        y = height - 40

        if data:
            keys = list(data[0].keys())
            for i, key in enumerate(keys):
                p.drawString(40 + i * 100, y, key)
            y -= 20
            for row in data:
                for i, value in enumerate(row.values()):
                    p.drawString(40 + i * 100, y, str(value))
                y -= 20
                if y < 40:
                    p.showPage()
                    y = height - 40
        else:
            p.drawString(40, y, "No data available.")

        p.save()
        buffer.seek(0)

        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename=appointments.pdf'
        return response


class AppointmentStatusUpdateView(UpdateAPIView):
    queryset = Appointment.objects.all()
    serializer_class = AppointmentSerializer
    permission_classes = [IsAuthenticated, IsAdmin | IsHealthcareProvider]
    lookup_url_kwarg = "appointment_id"

    def patch(self, request, *args, **kwargs):
        appointment = self.get_object()

        new_status = request.data.get("status")
        if new_status not in ['pending', 'completed', 'missed', 'cancelled']:
            return Response({"error": "Invalid status value."}, status=400)

        appointment.status = new_status
        appointment.save()
        ActivityLog.objects.create(
            actor=request.user,
            action="Updated appointment status",
            description=f"{request.user.username} updated status of appointment {appointment.id} to {new_status}"
        )

        return Response({"message": "Status updated successfully."}, status=drf_status.HTTP_200_OK)


class ActivityLogListView(generics.ListAPIView):
    queryset = ActivityLog.objects.select_related(
        'actor').order_by('-timestamp')
    serializer_class = ActivityLogSerializer
    permission_classes = [IsAuthenticated, IsAdmin]
    pagination_class = PageNumberPagination

    filter_backends = [DjangoFilterBackend, SearchFilter]
    # You can add more fields if needed
    filterset_class = ActivityLogFilter
    search_fields = ['action', 'description',
                     'actor__username', 'actor__first_name', 'actor__last_name']


class SystemSettingsView(APIView):
    permission_classes = [IsAuthenticated, IsAdmin]

    def get(self, request):
        settings, _ = SystemSettings.objects.get_or_create(id=1)
        serializer = SystemSettingsSerializer(settings)
        return Response(serializer.data)

    def put(self, request):
        settings, _ = SystemSettings.objects.get_or_create(id=1)
        serializer = SystemSettingsSerializer(settings, data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)


class UpdatePasswordView(APIView):
    permission_classes = [IsAuthenticated]

    def put(self, request):
        user = request.user
        old = request.data.get("old_password")
        new = request.data.get("new_password")

        if not old or not new:
            raise ValidationError({"detail": "Both fields are required."})

        if not user.check_password(old):
            raise ValidationError(
                {"old_password": "Incorrect current password."})

        # ✅ Validate new password strength using Django validators
        try:
            validate_password(new, user=user)
        except ValidationError as e:
            raise ValidationError({"new_password": e.messages})

        user.set_password(new)
        user.must_change_password = False  # ✅ Clear the flag after successful update
        user.save()

        return Response({"detail": "Password updated successfully."}, status=status.HTTP_200_OK)


class ProviderAddMotherView(generics.CreateAPIView):
    queryset = User.objects.filter(role='mother')
    serializer_class = ProviderAddMotherSerializer
    permission_classes = [IsAuthenticated, IsHealthcareProvider]

    def perform_create(self, serializer):
        user = serializer.save(created_by=self.request.user)

        # Generate custom_id and password
        user.custom_id = generate_custom_id('MOM')
        raw_password = generate_password()
        user.set_password(raw_password)
        user.save()

        # Log activity
        ActivityLog.objects.create(
            actor=self.request.user,
            action=f"Created mother: {user.username} ({user.custom_id})"
        )

        # Store credentials to attach in response
        self.generated_credentials = {
            "custom_id": user.custom_id,
            "username": user.username,
            "password": raw_password
        }

    def create(self, request, *args, **kwargs):
        super().create(request, *args, **kwargs)

        return Response(
            {
                "message": "Mother registered successfully.",
                "credentials": getattr(self, 'generated_credentials', {})
            },
            status=status.HTTP_201_CREATED
        )


# ✅ List Registered Mothers (paginated + searchable)
class ProviderListMothersView(generics.ListAPIView):
    serializer_class = MotherListSerializer
    permission_classes = [IsAuthenticated, IsHealthcareProvider]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_class = MotherFilter
    search_fields = ['custom_id', 'first_name', 'last_name', 'phone_number']
    pagination_class = PageNumberPagination

    def get_queryset(self):
        return User.objects.filter(role='mother', is_active=True)

# ✅ View Full Mother Details (for modal/sidebar)


class ProviderRetrieveMotherView(generics.RetrieveAPIView):
    queryset = User.objects.filter(role='mother', is_active=True)
    serializer_class = ProviderMotherDetailSerializer
    permission_classes = [IsAuthenticated, IsHealthcareProvider]
    lookup_field = 'custom_id'   # ✅ use custom_id since frontend sends MOM-xxxx


# ✅ Update Mother (already exists)


class ProviderUpdateMotherView(generics.UpdateAPIView):
    queryset = User.objects.filter(role='mother', is_active=True)
    serializer_class = ProviderUpdateMotherSerializer
    permission_classes = [IsAuthenticated, IsHealthcareProvider]
    lookup_field = 'custom_id'


# ✅ Export CSV


class ExportMothersCSVView(generics.ListAPIView):
    permission_classes = [IsAuthenticated, IsHealthcareProvider]

    def get(self, request, *args, **kwargs):
        mothers = User.objects.filter(role='mother')
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="registered_mothers.csv"'

        writer = csv.writer(response)
        writer.writerow(['Custom ID', 'Full Name', 'Phone', 'Date Created'])

        for mother in mothers:
            full_name = f"{mother.first_name} {mother.last_name}"
            writer.writerow([mother.custom_id, full_name,
                            mother.phone_number, mother.date_joined])

        return response

# ✅ Export PDF


class ExportMothersPDFView(generics.ListAPIView):
    permission_classes = [IsAuthenticated, IsHealthcareProvider]

    def get(self, request, *args, **kwargs):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="registered_mothers.pdf"'

        p = canvas.Canvas(response)
        y = 800

        mothers = User.objects.filter(role='mother')
        p.drawString(100, y, "Registered Mothers Report")
        y -= 40
        p.drawString(50, y, "ID")
        p.drawString(150, y, "Full Name")
        p.drawString(300, y, "Phone")
        p.drawString(400, y, "Date Created")

        for mother in mothers:
            y -= 20
            p.drawString(50, y, mother.custom_id)
            p.drawString(150, y, f"{mother.first_name} {mother.last_name}")
            p.drawString(300, y, mother.phone_number or "")
            p.drawString(400, y, str(mother.date_joined.strftime(
                "%Y-%m-%d")) if mother.date_joined else "")

        p.showPage()
        p.save()
        return response


class ProviderViewMothersView(generics.ListAPIView):
    serializer_class = MotherListSerializer
    permission_classes = [IsAuthenticated, IsHealthcareProvider]
    pagination_class = PageNumberPagination

    def get_queryset(self):
        queryset = User.objects.filter(role='mother', is_active=True)

        # Search filtering
        search_query = self.request.query_params.get('search')
        if search_query:
            queryset = queryset.filter(
                Q(first_name__icontains=search_query) |
                Q(last_name__icontains=search_query) |
                Q(custom_id__icontains=search_query) |
                Q(phone_number__icontains=search_query)
            )

        # Date range filtering
        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')

        if start_date and end_date:
            try:
                start = parse_date(start_date)
                end = parse_date(end_date)
                if start and end:
                    queryset = queryset.filter(
                        date_joined__date__range=[start, end])
            except:
                pass

        return queryset.order_by("-date_joined")


class ProviderDeleteMotherView(generics.DestroyAPIView):
    queryset = User.objects.filter(role='mother', is_active=True)
    serializer_class = ProviderUpdateMotherSerializer
    permission_classes = [IsAuthenticated, IsHealthcareProvider]
    lookup_field = 'custom_id'

    def perform_destroy(self, instance):
        # Soft delete instead of hard delete
        instance.is_active = False
        instance.save()

        ActivityLog.objects.create(
            actor=self.request.user,
            action="Soft-deleted mother",
            description=f"{self.request.user.username} soft-deleted mother {instance.custom_id}"
        )


# ========== Provider Views for Health Records =========

class ProviderMotherHealthView(APIView):
    permission_classes = [IsAuthenticated, IsHealthcareProvider]

    def _get_mother(self, custom_id):
        return get_object_or_404(User, custom_id=custom_id, role="mother", is_active=True)

    def get(self, request, custom_id):
        mother = self._get_mother(custom_id)
        record = PatientHealthRecord.objects.filter(
            mother=mother, is_active=True
        ).first()
        if not record:
            return Response(
                {"detail": "No health record found for this mother."},
                status=status.HTTP_404_NOT_FOUND,
            )
        return Response(PatientHealthRecordSerializer(record).data, status=status.HTTP_200_OK)

    def post(self, request, custom_id):
        mother = self._get_mother(custom_id)
        if PatientHealthRecord.objects.filter(mother=mother, is_active=True).exists():
            return Response(
                {"detail": "Health record already exists. Use PUT/PATCH to update."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        serializer = PatientHealthRecordSerializer(data=request.data)
        if serializer.is_valid():
            # ✅ auto-assign provider
            serializer.save(mother=mother, recorded_by=request.user)
            ActivityLog.objects.create(
                actor=request.user,
                action="Created health record",
                description=f"{request.user.username} created health record for {mother.custom_id}",
            )
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def put(self, request, custom_id):
        """Upsert behavior: create if missing, else update (full)."""
        mother = self._get_mother(custom_id)
        record, created = PatientHealthRecord.objects.get_or_create(
            mother=mother)
        if PatientHealthRecord.objects.filter(mother=mother, is_active=True).exclude(id=record.id).exists():
            return Response(
                {"detail": "Mother already has an active health record."},
                status=status.HTTP_400_BAD_REQUEST
            )
        serializer = PatientHealthRecordSerializer(
            record, data=request.data, partial=False)
        if serializer.is_valid():
            # ✅ ensure provider is recorded
            serializer.save(recorded_by=request.user)
            ActivityLog.objects.create(
                actor=request.user,
                action="Created health record" if created else "Updated health record",
                description=f"{request.user.username} {'created' if created else 'updated'} health record for {mother.custom_id}",
            )
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def patch(self, request, custom_id):
        mother = self._get_mother(custom_id)
        record = get_object_or_404(PatientHealthRecord, mother=mother)
        serializer = PatientHealthRecordSerializer(
            record, data=request.data, partial=True)
        if serializer.is_valid():
            # ✅ always track last updater
            serializer.save(recorded_by=request.user)
            ActivityLog.objects.create(
                actor=request.user,
                action="Updated health record",
                description=f"{request.user.username} patched health record for {mother.custom_id}",
            )
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class ProviderExaminationListCreateView(generics.ListCreateAPIView):
    serializer_class = ExaminationRecordSerializer
    permission_classes = [IsAuthenticated, IsHealthcareProvider]

    def _get_mother(self):
        return get_object_or_404(User, custom_id=self.kwargs["custom_id"], role="mother", is_active=True)

    def get_queryset(self):
        mother = self._get_mother()
        return ExaminationRecord.objects.filter(mother=mother, is_active=True)

    def perform_create(self, serializer):
        mother = self._get_mother()
        instance = serializer.save(mother=mother, provider=self.request.user)
        ActivityLog.objects.create(
            actor=self.request.user,
            action="Created exam record",
            description=f"{self.request.user.username} added exam for {mother.custom_id} (id={instance.id})",
        )


class ProviderExaminationDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = ExaminationRecordSerializer
    permission_classes = [IsAuthenticated, IsHealthcareProvider]

    def _get_mother(self):
        return get_object_or_404(User, custom_id=self.kwargs["custom_id"], role="mother", is_active=True)

    def get_queryset(self):
        mother = self._get_mother()
        return ExaminationRecord.objects.filter(mother=mother, is_active=True)

    def perform_update(self, serializer):
        instance = serializer.save()
        ActivityLog.objects.create(
            actor=self.request.user,
            action="Updated exam record",
            description=f"{self.request.user.username} updated exam #{instance.id} for {instance.mother.custom_id}",
        )

    def perform_destroy(self, instance):
        instance.is_active = False
        instance.save(update_fields=["is_active"])
        ActivityLog.objects.create(
            actor=self.request.user,
            action="Soft-deleted exam record",
            description=f"{self.request.user.username} soft-deleted exam #{instance.id} for {instance.mother.custom_id}",
        )


class ProviderPreviousPregnancyListCreateView(generics.ListCreateAPIView):
    serializer_class = PreviousPregnancySerializer
    permission_classes = [IsAuthenticated, IsHealthcareProvider]

    def _get_record(self):
        mother = get_object_or_404(
            User, custom_id=self.kwargs["custom_id"], role="mother", is_active=True)
        return get_object_or_404(PatientHealthRecord, mother=mother, is_active=True)

    def get_queryset(self):
        return PreviousPregnancy.objects.filter(health_record=self._get_record(),
                                                is_active=True)

    def perform_create(self, serializer):
        record = self._get_record()
        instance = serializer.save(health_record=record)
        ActivityLog.objects.create(
            actor=self.request.user,
            action="Created previous pregnancy",
            description=f"{self.request.user.username} added prev pregnancy (id={instance.id}) for {record.mother.custom_id}",
        )


class ProviderPreviousPregnancyDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = PreviousPregnancySerializer
    permission_classes = [IsAuthenticated, IsHealthcareProvider]

    def _get_record(self):
        mother = get_object_or_404(
            User, custom_id=self.kwargs["custom_id"], role="mother", is_active=True)
        return get_object_or_404(PatientHealthRecord, mother=mother, is_active=True)

    def get_queryset(self):
        return PreviousPregnancy.objects.filter(health_record=self._get_record())

    def perform_update(self, serializer):
        instance = serializer.save()
        ActivityLog.objects.create(
            actor=self.request.user,
            action="Updated previous pregnancy",
            description=f"{self.request.user.username} updated prev pregnancy #{instance.id} for {instance.health_record.mother.custom_id}",
        )

    def perform_destroy(self, instance):
        instance.is_active = False
        instance.save(update_fields=["is_active"])
        ActivityLog.objects.create(
            actor=self.request.user,
            action="Soft-deleted previous pregnancy",
            description=f"{self.request.user.username} soft-deleted prev pregnancy #{instance.id} for {instance.health_record.mother.custom_id}",
        )


# ========== Mother Views for Health Records (Read-Only) ==========

class MotherHealthRecordView(APIView):
    permission_classes = [IsAuthenticated, IsMother]

    def get(self, request):
        record = (
            PatientHealthRecord.objects
            .filter(mother=request.user, is_active=True)
            .select_related("mother")
            .prefetch_related(
                "previous_pregnancies",  # already correct
                Prefetch(
                    "mother__examinations",
                    queryset=ExaminationRecord.objects.filter(
                        is_active=True).order_by("-visit_date", "-id")
                ),
            )
        ).first()

        if not record:
            return Response({"detail": "No health record found."}, status=status.HTTP_404_NOT_FOUND)

        serializer = PatientHealthRecordSerializer(record)
        return Response(serializer.data, status=status.HTTP_200_OK)


class MotherExaminationListView(generics.ListAPIView):
    serializer_class = ExaminationRecordSerializer
    permission_classes = [IsAuthenticated, IsMother]

    def get_queryset(self):
        return ExaminationRecord.objects.filter(
            mother=self.request.user,
            is_active=True
        ).order_by("-visit_date")


class MotherPreviousPregnancyListView(generics.ListAPIView):
    serializer_class = PreviousPregnancySerializer
    permission_classes = [IsAuthenticated, IsMother]

    def get_queryset(self):
        return PreviousPregnancy.objects.filter(
            health_record__mother=self.request.user,
            health_record__is_active=True,
            is_active=True  # ✅ only active pregnancies
        ).order_by("-id")


# Export Mother Health Record
class ExportMotherHealthView(APIView):
    permission_classes = [IsAuthenticated, IsHealthcareProvider]

    def get(self, request, custom_id):
        fmt = (request.GET.get("format") or "pdf").lower()
        mother = get_object_or_404(
            User, custom_id=custom_id, role="mother", is_active=True)

        health = get_object_or_404(
            PatientHealthRecord, mother=mother, is_active=True)
        prev_pregs = PreviousPregnancy.objects.filter(
            health_record=health, is_active=True).order_by("id")
        visits = ExaminationRecord.objects.filter(
            mother=mother, is_active=True).order_by("-visit_date", "-id")

        filename_base = f"{mother.custom_id or mother.id}_health_record"

        if fmt == "pdf":
            buffer = self._build_pdf(mother, health, prev_pregs, visits)
            resp = HttpResponse(buffer.getvalue(),
                                content_type="application/pdf")
            resp["Content-Disposition"] = f'attachment; filename="{filename_base}.pdf"'
            return resp

        if fmt in ("xlsx", "excel"):
            buffer = self._build_xlsx(mother, health, prev_pregs, visits)
            resp = HttpResponse(
                buffer.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            resp["Content-Disposition"] = f'attachment; filename="{filename_base}.xlsx"'
            return resp

        if fmt == "csv":
            buffer = self._build_csv(mother, health, prev_pregs, visits)
            resp = HttpResponse(buffer.getvalue(),
                                content_type="text/csv; charset=utf-8")
            resp["Content-Disposition"] = f'attachment; filename="{filename_base}.csv"'
            return resp

        return HttpResponse(f"Unsupported format: {fmt}", status=400)

    # =========================
    # Helpers: PDF / XLSX / CSV
    # =========================

    def _build_pdf(self, mother, health, prev_pregs, visits):
        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf, pagesize=A4, title="Maternal Health Record")
        styles = getSampleStyleSheet()
        H1 = styles["Heading1"]
        H2 = styles["Heading2"]
        P = styles["BodyText"]

        story = []

        # Header
        story.append(Paragraph("Maternal Health Record", H1))
        story.append(
            Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", P))
        story.append(Spacer(1, 10))

        # Mother bio
        story.append(Paragraph("Patient", H2))
        bio_rows = [
            ["ID", mother.custom_id or ""],
            ["Name", f"{mother.first_name or ''} {mother.last_name or ''}".strip(
            ) or mother.username],
            ["Phone", mother.phone_number or ""],
            ["Address", mother.address or ""],
            ["DOB", mother.date_of_birth or ""],
        ]
        story.append(self._table(bio_rows))

        # Health record core
        story.append(Spacer(1, 10))
        story.append(Paragraph("Health Record (Baseline)", H2))
        hr_rows = [
            ["Blood Group", health.blood_group or "-",
                "Genotype", health.genotype or "-"],
            ["Height (cm)", health.height_cm or "-",
             "Gravidity", health.gravidity or "-"],
            ["Parity", health.parity or "-", "LMP", health.lmp or "-"],
            ["EDD", health.edd or "-", "Infertility",
                health.infertility_status or "-"],
            ["Allergies", health.allergies or "-",
                "Chronic Conditions", health.chronic_conditions or "-"],
            ["Medications", health.medications or "-", "Family Planning",
                health.recent_family_planning_method or "-"],
            ["Prev Illness", health.previous_illness or "-",
                "Prev Surgery", health.previous_surgery or "-"],
            ["Family History", health.family_history or "-", "", ""],
        ]
        story.append(self._table(hr_rows, colWidths=[90, 180, 110, 150]))

        # Investigations
        story.append(Spacer(1, 10))
        story.append(Paragraph("Investigations", H2))
        inv_rows = [
            ["Father Blood Group", health.blood_group_father or "-",
                "Mother Rh", health.rhesus_factor_mother or "-"],
            ["Father Rh", health.rhesus_factor_father or "-",
                "Hep B", health.hepatitis_b_status or "-"],
            ["VDRL", health.vdrl_status or "-", "RV", health.rv_status or "-"],
            ["Hb (Booking)", health.haemoglobin_booking or "-",
             "Hb (28w)", health.haemoglobin_28w or "-"],
            ["Hb (36w)", health.haemoglobin_36w or "-", "", ""],
            ["US 1 Date", health.ultrasound1_date or "-",
                "US 1 Result", (health.ultrasound1_result or "")[:80]],
            ["US 2 Date", health.ultrasound2_date or "-",
                "US 2 Result", (health.ultrasound2_result or "")[:80]],
            ["Pap Smear Date", health.pap_smear_date or "-",
                "Pap Smear Comment", (health.pap_smear_comments or "")[:80]],
        ]
        story.append(self._table(inv_rows, colWidths=[100, 150, 90, 190]))

        # Previous Pregnancies
        story.append(Spacer(1, 10))
        story.append(Paragraph("Previous Pregnancies", H2))
        if prev_pregs:
            preg_rows = [["Place of Birth", "Gestation", "Mode",
                          "Labour", "Outcome", "Birth Wt", "Complications"]]
            for p in prev_pregs:
                preg_rows.append([
                    p.place_of_birth or "-",
                    p.gestation_at_delivery or "-",
                    p.mode_of_delivery or "-",
                    p.labour_duration or "-",
                    p.outcome or "-",
                    p.birth_weight or "-",
                    (p.complications or "")[:40],
                ])
            story.append(self._table(preg_rows, header=True))
        else:
            story.append(Paragraph("None recorded.", P))

        # Examinations (visits)
        story.append(Spacer(1, 10))
        story.append(Paragraph("Antenatal Visits / Examinations", H2))
        if visits:
            visit_rows = [[
                "Date", "GA(w)", "BP", "Wt(kg)", "Temp", "Pulse", "Resp",
                "Fundal Ht", "FHR", "Ur Prot", "Ur Glu", "Oedema", "Pres.", "Lie",
                "Next Appt", "Provider", "Notes"
            ]]
            for v in visits:
                provider_name = ""
                if v.provider:
                    fn = (v.provider.first_name or "").strip()
                    ln = (v.provider.last_name or "").strip()
                    provider_name = (f"{fn} {ln}".strip() or getattr(
                        v.provider, "full_name", "") or v.provider.username)

                visit_rows.append([
                    (v.visit_date or "").strftime(
                        "%Y-%m-%d") if v.visit_date else "-",
                    v.gestational_age_weeks or "-",
                    f"{v.blood_pressure_systolic or '-'} / {v.blood_pressure_diastolic or '-'}",
                    v.weight_kg or "-", v.temperature_c or "-", v.pulse_rate or "-", v.respiratory_rate or "-",
                    v.fundal_height_cm or "-", v.fetal_heart_rate or "-",
                    v.urine_protein or "-", v.urine_glucose or "-",
                    v.oedema or "-", v.presentation or "-", v.lie or "-",
                    v.next_appointment or "-", provider_name or "-",
                    (v.notes or "")[:60],
                ])
            story.append(self._table(visit_rows, header=True))
        else:
            story.append(Paragraph("No visits recorded.", P))

        doc.build(story)
        buf.seek(0)
        return buf

    def _table(self, rows, header=False, colWidths=None):
        t = Table(rows, colWidths=colWidths)
        style = TableStyle([
            ("BOX", (0, 0), (-1, -1), 0.25, colors.grey),
            ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("FONT", (0, 0), (-1, -1), "Helvetica", 9),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ])
        if header:
            style.add("BACKGROUND", (0, 0), (-1, 0), colors.whitesmoke)
            style.add("FONT", (0, 0), (-1, 0), "Helvetica-Bold", 9)
        t.setStyle(style)
        return t

    def _build_xlsx(self, mother, health, prev_pregs, visits):
        wb = Workbook()

        # Sheet 1: Patient + Baseline
        ws = wb.active
        ws.title = "Patient & Baseline"

        ws.append(["Patient"])
        ws.append(["ID", mother.custom_id or ""])
        full_name = f"{mother.first_name or ''} {mother.last_name or ''}".strip(
        ) or mother.username
        ws.append(["Name", full_name])
        ws.append(["Phone", mother.phone_number or ""])
        ws.append(["Address", mother.address or ""])
        ws.append(["DOB", str(mother.date_of_birth or "")])

        ws.append([])
        ws.append(["Health Record (Baseline)"])
        baseline = [
            ["Blood Group", health.blood_group or "",
                "Genotype", health.genotype or ""],
            ["Height (cm)", health.height_cm or "",
             "Gravidity", health.gravidity or ""],
            ["Parity", health.parity or "", "LMP", str(health.lmp or "")],
            ["EDD", str(health.edd or ""), "Infertility",
             health.infertility_status or ""],
            ["Allergies", health.allergies or "",
                "Chronic Conditions", health.chronic_conditions or ""],
            ["Medications", health.medications or "", "Family Planning",
                health.recent_family_planning_method or ""],
            ["Prev Illness", health.previous_illness or "",
                "Prev Surgery", health.previous_surgery or ""],
            ["Family History", health.family_history or "", "", ""],
        ]
        for r in baseline:
            ws.append(r)

        ws.append([])
        ws.append(["Investigations"])
        inv = [
            ["Father Blood Group", health.blood_group_father or "",
                "Mother Rh", health.rhesus_factor_mother or ""],
            ["Father Rh", health.rhesus_factor_father or "",
                "Hep B", health.hepatitis_b_status or ""],
            ["VDRL", health.vdrl_status or "", "RV", health.rv_status or ""],
            ["Hb (Booking)", health.haemoglobin_booking or "",
             "Hb (28w)", health.haemoglobin_28w or ""],
            ["Hb (36w)", health.haemoglobin_36w or "", "", ""],
            ["US 1 Date", str(health.ultrasound1_date or ""),
             "US 1 Result", health.ultrasound1_result or ""],
            ["US 2 Date", str(health.ultrasound2_date or ""),
             "US 2 Result", health.ultrasound2_result or ""],
            ["Pap Smear Date", str(health.pap_smear_date or ""),
             "Pap Smear Comment", health.pap_smear_comments or ""],
        ]
        for r in inv:
            ws.append(r)

        # Auto-width-ish
        for col in ws.columns:
            max_len = max(len(str(c.value))
                          if c.value is not None else 0 for c in col)
            ws.column_dimensions[get_column_letter(
                col[0].column)].width = min(max(12, max_len + 2), 50)

        # Sheet 2: Previous Pregnancies
        ws2 = wb.create_sheet(title="Previous Pregnancies")
        ws2.append(["Place of Birth", "Gestation", "Mode", "Labour",
                   "Outcome", "Birth Weight", "Complications"])
        for p in prev_pregs:
            ws2.append([
                p.place_of_birth or "",
                p.gestation_at_delivery or "",
                p.mode_of_delivery or "",
                p.labour_duration or "",
                p.outcome or "",
                p.birth_weight or "",
                p.complications or "",
            ])

        for col in ws2.columns:
            max_len = max(len(str(c.value))
                          if c.value is not None else 0 for c in col)
            ws2.column_dimensions[get_column_letter(
                col[0].column)].width = min(max(12, max_len + 2), 50)

        # Sheet 3: Examinations
        ws3 = wb.create_sheet(title="Examinations")
        ws3.append([
            "Visit Date", "GA (weeks)", "BP Systolic", "BP Diastolic", "Weight (kg)", "Temp (°C)",
            "Pulse", "Resp", "Fundal Ht (cm)", "FHR", "Urine Protein", "Urine Glucose",
            "Oedema", "Presentation", "Lie", "Next Appointment", "Provider", "Notes"
        ])
        for v in visits:
            provider_name = ""
            if v.provider:
                fn = (v.provider.first_name or "").strip()
                ln = (v.provider.last_name or "").strip()
                provider_name = (f"{fn} {ln}".strip() or getattr(
                    v.provider, "full_name", "") or v.provider.username)

            ws3.append([
                v.visit_date.strftime("%Y-%m-%d") if v.visit_date else "",
                v.gestational_age_weeks or "",
                v.blood_pressure_systolic or "",
                v.blood_pressure_diastolic or "",
                v.weight_kg or "",
                v.temperature_c or "",
                v.pulse_rate or "",
                v.respiratory_rate or "",
                v.fundal_height_cm or "",
                v.fetal_heart_rate or "",
                v.urine_protein or "",
                v.urine_glucose or "",
                v.oedema or "",
                v.presentation or "",
                v.lie or "",
                v.next_appointment or "",
                provider_name,
                v.notes or "",
            ])

        for col in ws3.columns:
            max_len = max(len(str(c.value))
                          if c.value is not None else 0 for c in col)
            ws3.column_dimensions[get_column_letter(
                col[0].column)].width = min(max(12, max_len + 2), 50)

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return out

    def _build_csv(self, mother, health, prev_pregs, visits):
        # Single CSV with simple section markers.
        s = io.StringIO()
        w = csv.writer(s)

        w.writerow(["Maternal Health Record"])
        w.writerow(["Generated", datetime.now().strftime("%Y-%m-%d %H:%M")])
        w.writerow([])

        # Patient
        w.writerow(["# Patient"])
        w.writerow(["ID", mother.custom_id or ""])
        full_name = f"{mother.first_name or ''} {mother.last_name or ''}".strip(
        ) or mother.username
        w.writerow(["Name", full_name])
        w.writerow(["Phone", mother.phone_number or ""])
        w.writerow(["Address", mother.address or ""])
        w.writerow(["DOB", mother.date_of_birth or ""])
        w.writerow([])

        # Baseline
        w.writerow(["# Health Record (Baseline)"])
        baseline = [
            ["Blood Group", health.blood_group or "",
                "Genotype", health.genotype or ""],
            ["Height (cm)", health.height_cm or "",
             "Gravidity", health.gravidity or ""],
            ["Parity", health.parity or "", "LMP", health.lmp or ""],
            ["EDD", health.edd or "", "Infertility",
                health.infertility_status or ""],
            ["Allergies", health.allergies or "",
                "Chronic Conditions", health.chronic_conditions or ""],
            ["Medications", health.medications or "", "Family Planning",
                health.recent_family_planning_method or ""],
            ["Prev Illness", health.previous_illness or "",
                "Prev Surgery", health.previous_surgery or ""],
            ["Family History", health.family_history or "", "", ""],
        ]
        w.writerows(baseline)
        w.writerow([])

        # Investigations
        w.writerow(["# Investigations"])
        inv = [
            ["Father Blood Group", health.blood_group_father or "",
                "Mother Rh", health.rhesus_factor_mother or ""],
            ["Father Rh", health.rhesus_factor_father or "",
                "Hep B", health.hepatitis_b_status or ""],
            ["VDRL", health.vdrl_status or "", "RV", health.rv_status or ""],
            ["Hb (Booking)", health.haemoglobin_booking or "",
             "Hb (28w)", health.haemoglobin_28w or ""],
            ["Hb (36w)", health.haemoglobin_36w or "", "", ""],
            ["US 1 Date", health.ultrasound1_date or "",
                "US 1 Result", health.ultrasound1_result or ""],
            ["US 2 Date", health.ultrasound2_date or "",
                "US 2 Result", health.ultrasound2_result or ""],
            ["Pap Smear Date", health.pap_smear_date or "",
                "Pap Smear Comment", health.pap_smear_comments or ""],
        ]
        w.writerows(inv)
        w.writerow([])

        # Previous Pregnancies
        w.writerow(["# Previous Pregnancies"])
        w.writerow(["Place of Birth", "Gestation", "Mode", "Labour",
                   "Outcome", "Birth Weight", "Complications"])
        for p in prev_pregs:
            w.writerow([
                p.place_of_birth or "",
                p.gestation_at_delivery or "",
                p.mode_of_delivery or "",
                p.labour_duration or "",
                p.outcome or "",
                p.birth_weight or "",
                p.complications or "",
            ])
        w.writerow([])

        # Examinations
        w.writerow(["# Examinations"])
        w.writerow([
            "Visit Date", "GA (weeks)", "BP Systolic", "BP Diastolic",
            "Weight (kg)", "Temp (°C)", "Pulse", "Resp", "Fundal Ht (cm)",
            "FHR", "Urine Protein", "Urine Glucose", "Oedema", "Presentation",
            "Lie", "Next Appointment", "Provider", "Notes"
        ])
        for v in visits:
            provider_name = ""
            if v.provider:
                fn = (v.provider.first_name or "").strip()
                ln = (v.provider.last_name or "").strip()
                provider_name = (f"{fn} {ln}".strip() or getattr(
                    v.provider, "full_name", "") or v.provider.username)

            w.writerow([
                v.visit_date.strftime("%Y-%m-%d") if v.visit_date else "",
                v.gestational_age_weeks or "",
                v.blood_pressure_systolic or "",
                v.blood_pressure_diastolic or "",
                v.weight_kg or "",
                v.temperature_c or "",
                v.pulse_rate or "",
                v.respiratory_rate or "",
                v.fundal_height_cm or "",
                v.fetal_heart_rate or "",
                v.urine_protein or "",
                v.urine_glucose or "",
                v.oedema or "",
                v.presentation or "",
                v.lie or "",
                v.next_appointment or "",
                provider_name,
                v.notes or "",
            ])

        out = io.StringIO()
        out.write(s.getvalue())
        out.seek(0)
        return out

# Mother Upcoming Appointments


class MotherUpcomingAppointments(generics.ListAPIView):
    permission_classes = [IsAuthenticated, IsMother]
    serializer_class = ExaminationRecordSerializer

    def get_queryset(self):
        return (ExaminationRecord.objects
                .filter(mother=self.request.user,
                        is_active=True,
                        next_appointment__gte=timezone.localdate())
                .order_by("next_appointment"))


# class MotherNotificationListView(generics.ListAPIView):
#     permission_classes = [IsAuthenticated, IsMother]
#     serializer_class = NotificationSerializer

#     def get_queryset(self):
#         qs = Notification.objects.filter(user=self.request.user)
#         if self.request.query_params.get("unread") == "true":
#             qs = qs.filter(read=False)
#         return qs.order_by("-created_at")


class NotificationMarkReadView(APIView):
    permission_classes = [IsAuthenticated, IsMother]

    def post(self, request, pk):
        notif = get_object_or_404(Notification, pk=pk, user=request.user)
        notif.read = True
        notif.save(update_fields=["read"])
        return Response({"status": "ok"})

# Provider Dashboard Metrics


class ProviderDashboardMetricsView(APIView):
    permission_classes = [IsAuthenticated, IsHealthcareProvider]

    def get(self, request):
        today = timezone.localdate()
        week_end = today + timedelta(days=7)

        total_mothers = User.objects.filter(
            role="mother", is_active=True).count()

        upcoming_today = ExaminationRecord.objects.filter(
            next_appointment=today, is_active=True
        ).count()

        upcoming_week = ExaminationRecord.objects.filter(
            next_appointment__gt=today,
            next_appointment__lte=week_end,
            is_active=True,
        ).count()

        missed = ExaminationRecord.objects.filter(
            next_appointment__lt=today, is_active=True
        ).count()

        return Response({
            "total_mothers": total_mothers,
            "upcoming_today": upcoming_today,
            "upcoming_week": upcoming_week,
            "missed": missed,
        })


class ProviderUpcomingAppointmentsView(ListAPIView):
    serializer_class = ExaminationRecordSerializer
    permission_classes = [IsAuthenticated, IsHealthcareProvider]

    def get_queryset(self):
        today = timezone.localdate()
        return ExaminationRecord.objects.filter(
            next_appointment__gte=today, is_active=True
        ).select_related("mother", "provider").order_by("next_appointment")[:10]


class ProviderRecentVisitsView(ListAPIView):
    serializer_class = ExaminationRecordSerializer
    permission_classes = [IsAuthenticated, IsHealthcareProvider]

    def get_queryset(self):
        return ExaminationRecord.objects.filter(is_active=True).select_related("mother", "provider").order_by("-created_at")[:6]


class ProviderRecentPatientsView(ListAPIView):
    serializer_class = UserSerializer
    permission_classes = [IsAuthenticated, IsHealthcareProvider]

    def get_queryset(self):
        return User.objects.filter(role="mother", is_active=True).order_by("-date_joined")[:6]


class ProviderReminderLogView(generics.ListAPIView):
    serializer_class = AppointmentReminderSerializer
    permission_classes = [IsAuthenticated, IsHealthcareProvider]

    def get_queryset(self):
        return AppointmentReminder.objects.select_related("exam", "exam__mother").order_by("-sent_at")[:6]


class NotificationListView(generics.ListAPIView):
    """
    Notifications for the current authenticated user.
    Query param: ?unread=true  -> returns only unread notifications
    """
    permission_classes = [IsAuthenticated]
    serializer_class = NotificationSerializer

    def get_queryset(self):
        qs = Notification.objects.filter(user=self.request.user)
        unread = self.request.query_params.get("unread")
        if unread == "true":
            qs = qs.filter(read=False)
        return qs.order_by("-created_at")
# ------------------------------------------------------------------------

# ------------------------------------------------------
# Get Current User Information


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_me(request):
    user = request.user
    return Response({
        "id": user.id,
        "name": user.get_full_name() or user.username,
        "email": user.email,
        "role": user.role,
    })

# User Logout


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def logout_view(request):
    try:
        refresh_token = request.data.get("refresh")
        token = RefreshToken(refresh_token)
        token.blacklist()
        return Response({"message": "Successfully logged out."})
    except Exception:
        return Response({"error": "Invalid token."}, status=400)


class MotherProfileView(generics.RetrieveAPIView):
    serializer_class = MotherProfileSerializer
    permission_classes = [IsAuthenticated, IsMother]

    def get_object(self):
        return self.request.user


class MotherProfileUpdateView(generics.UpdateAPIView):
    serializer_class = MotherContactUpdateSerializer
    permission_classes = [IsAuthenticated, IsMother]

    def get_object(self):
        return self.request.user


class MotherAppointmentsListView(generics.ListAPIView):
    serializer_class = AppointmentSerializer
    permission_classes = [IsAuthenticated, IsMother]

    def get_queryset(self):
        return Appointment.objects.filter(
            patient=self.request.user,
            status="upcoming"   # ✅ instead of is_active
        ).order_by("appointment_date")  # soonest first


class MotherChangePasswordView(UpdateAPIView):
    serializer_class = ChangePasswordSerializer
    permission_classes = [IsAuthenticated]  # or [IsAuthenticated, IsMother]

    def get_object(self):
        return self.request.user

    def update(self, request, *args, **kwargs):
        user = self.get_object()
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        if not user.check_password(serializer.validated_data['old_password']):
            raise ValidationError({'old_password': 'Incorrect old password.'})

        user.set_password(serializer.validated_data['new_password'])
        user.save()

        return Response({'detail': 'Password updated successfully.'}, status=status.HTTP_200_OK)


class MotherNotificationListView(generics.ListAPIView):
    serializer_class = NotificationSerializer
    permission_classes = [IsAuthenticated, IsMother]

    def get_queryset(self):
        return Notification.objects.filter(
            user=self.request.user
        ).order_by("-created_at")


class MotherNotificationMarkReadView(APIView):
    permission_classes = [IsAuthenticated, IsMother]

    def post(self, request, pk=None):
        if pk:
            notif = get_object_or_404(Notification, id=pk, user=request.user)
            notif.read = True
            notif.save()
            return Response({"detail": "Notification marked as read."})
        else:
            Notification.objects.filter(
                user=request.user, read=False).update(read=True)
            return Response({"detail": "All notifications marked as read."})


class MotherEmergencyAlertView(APIView):
    permission_classes = [IsAuthenticated, IsMother]

    def post(self, request):
        message = request.data.get("message", "Emergency alert triggered")

        # Notify all admins and providers
        recipients = User.objects.filter(
            role__in=["admin", "provider"], is_active=True)

        for recipient in recipients:
            Notification.objects.create(
                user=recipient,
                title="🚨 Emergency Alert",
                body=f"{request.user.get_full_name()} ({request.user.custom_id}) "
                     f"triggered an emergency: {message}"
            )

        return Response(
            {"detail": "Emergency alert sent."},
            status=status.HTTP_201_CREATED
        )


# from .permissions import IsMother   # if you have this


class MotherDashboardView(APIView):
    permission_classes = [IsAuthenticated]  # or [IsAuthenticated, IsMother]

    def get(self, request):
        user = request.user
        if getattr(user, "role", None) != "mother":
            return Response({"detail": "Unauthorized"}, status=403)

        today = timezone.localdate()

        # Upcoming appointment from ExaminationRecord.next_appointment (DateField)
        next_exam = (
            ExaminationRecord.objects
            .select_related("provider")
            .filter(mother=user, is_active=True, next_appointment__gte=today)
            .order_by("next_appointment")
            .first()
        )

        provider_name = None
        if next_exam and next_exam.provider:
            fn = (next_exam.provider.first_name or "").strip()
            ln = (next_exam.provider.last_name or "").strip()
            provider_name = (f"{fn} {ln}".strip()
                             or next_exam.provider.username)

        next_appointment = None
        if next_exam and next_exam.next_appointment:
            next_appointment = {
                "id": next_exam.id,
                "date": next_exam.next_appointment,   # ISO date from DRF
                "provider_name": provider_name,
                "notes": next_exam.notes or "",
            }

        # Latest exam for health overview + risk
        latest_exam = (
            ExaminationRecord.objects
            .filter(mother=user, is_active=True)
            .order_by("-visit_date")
            .first()
        )

        # Notifications (last 5)
        notifications = list(
            Notification.objects
            .filter(user=user)
            .order_by("-created_at")
            .values("id", "title", "body", "created_at")[:5]
        )

        # Health summary
        full_name = f"{(user.first_name or '').strip()} {(user.last_name or '').strip()}".strip(
        ) or user.username
        health_summary = {
            "name": full_name,
            "phone": user.phone_number,     # <- your field
            "address": user.address,
            "last_visit": latest_exam.visit_date if latest_exam else None,
            "pregnancy_week": latest_exam.gestational_age_weeks if latest_exam else None,
            "risk_status": latest_exam.assess_risk() if latest_exam else "Unknown",
        }

        return Response({
            "next_appointment": next_appointment,
            "notifications": notifications,
            "health_summary": health_summary,
        })


# ------------------ ADMIN DASHBOARD ------------------
class AdminDashboardView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        if user.role != "admin":
            return Response({"detail": "Unauthorized"}, status=403)

        data = {
            "total_mothers": User.objects.filter(role="mother").count(),
            "total_providers": User.objects.filter(role="provider").count(),
            "total_appointments": Appointment.objects.count(),
            "total_examinations": ExaminationRecord.objects.count() if ExaminationRecord else 0,
        }
        return Response(data)


# ------------------ PROVIDER DASHBOARD ------------------
class ProviderDashboardView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        if user.role != "provider":
            return Response({"detail": "Unauthorized"}, status=403)

        data = {
            # or filter by provider if needed
            "my_patients": User.objects.filter(role="mother").count(),
            "my_appointments": Appointment.objects.filter(provider=user).count(),
            "upcoming_appointments": Appointment.objects.filter(
                provider=user, date__gte=timezone.now()
            ).order_by("date")[:5].values("id", "date", "mother__full_name"),
        }
        return Response(data)


class MotherNextAppointmentView(generics.RetrieveAPIView):
    serializer_class = AppointmentSerializer
    permission_classes = [IsAuthenticated, IsMother]

    def get_object(self):
        return Appointment.objects.filter(
            patient=self.request.user,
            status="pending",
            appointment_date__gte=timezone.now()
        ).order_by("appointment_date").first()


class MotherAllAppointmentsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        today = timezone.localdate()

        # 1️⃣ Admin-created appointments
        admin_appts = Appointment.objects.filter(patient=user)

        # 2️⃣ Provider-created with next appointment
        exam_appts = ExaminationRecord.objects.filter(
            mother=user,
            next_appointment__isnull=False
        )

        # 3️⃣ Serialize using the same unified serializer
        data = []
        for a in admin_appts:
            data.append(UnifiedAppointmentSerializer.from_admin(a))
        for e in exam_appts:
            data.append(UnifiedAppointmentSerializer.from_exam(e))

        # 4️⃣ Split into upcoming / past
        upcoming = []
        past = []
        for item in data:
            appt_date = item["appointment_date"].date() if hasattr(
                item["appointment_date"], "date") else item["appointment_date"]
            if appt_date >= today:
                upcoming.append(item)
            else:
                past.append(item)

        return Response({
            "upcoming": upcoming,
            "past": past
        })


class ProviderNotificationMarkReadView(APIView):
    permission_classes = [IsAuthenticated]

    def patch(self, request, pk=None):
        try:
            notif = Notification.objects.get(pk=pk, user=request.user)
            notif.read = True
            notif.save()
            return Response(
                {"detail": "Notification marked as read",
                 "notification": NotificationSerializer(notif).data},
                status=status.HTTP_200_OK
            )
        except Notification.DoesNotExist:
            return Response(
                {"detail": "Notification not found"},
                status=status.HTTP_404_NOT_FOUND
            )


class ProviderNotificationMarkAllReadView(APIView):
    permission_classes = [IsAuthenticated]

    def patch(self, request):
        updated = Notification.objects.filter(
            user=request.user, read=False
        ).update(read=True)

        return Response(
            {"detail": f"{updated} notifications marked as read"},
            status=status.HTTP_200_OK
        )


class ProviderProfileView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if request.user.role != "provider":
            return Response({"detail": "Not allowed"}, status=status.HTTP_403_FORBIDDEN)
        serializer = UserSerializer(request.user)
        return Response(serializer.data)

    def put(self, request):
        if request.user.role != "provider":
            return Response({"detail": "Not allowed"}, status=status.HTTP_403_FORBIDDEN)
        serializer = UserSerializer(
            request.user, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class ProviderChangePasswordView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if request.user.role != "provider":
            return Response({"detail": "Not allowed"}, status=status.HTTP_403_FORBIDDEN)

        old_password = request.data.get("old_password")
        new_password = request.data.get("new_password")
        confirm_password = request.data.get("confirm_password")

        if not all([old_password, new_password, confirm_password]):
            return Response({"detail": "All fields are required"}, status=status.HTTP_400_BAD_REQUEST)

        if not check_password(old_password, request.user.password):
            return Response({"detail": "Old password is incorrect"}, status=status.HTTP_400_BAD_REQUEST)

        if new_password != confirm_password:
            return Response({"detail": "Passwords do not match"}, status=status.HTTP_400_BAD_REQUEST)

        request.user.set_password(new_password)
        request.user.save()
        return Response({"detail": "Password updated successfully"})


class ProviderProfileUpdateView(APIView):
    permission_classes = [IsAuthenticated]

    def put(self, request):
        serializer = ProviderProfileSerializer(
            request.user, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class ProviderDashboardView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        provider = request.user

        # Ensure only providers can see this
        if provider.role != "provider":
            return Response({"detail": "Not authorized"}, status=403)

        # Metrics
        total_mothers = User.objects.filter(
            role="mother", created_by=provider).count()
        total_appointments = Appointment.objects.filter(
            provider=provider).count()
        upcoming_appointments_count = Appointment.objects.filter(
            provider=provider, status="pending", appointment_date__gte=now()
        ).count()
        completed_visits = ExaminationRecord.objects.filter(
            provider=provider).count()

        metrics = {
            "total_mothers": total_mothers,
            "total_appointments": total_appointments,
            "upcoming_appointments": upcoming_appointments_count,
            "completed_visits": completed_visits,
        }

        # Appointment status distribution
        appointments_by_status = (
            Appointment.objects.filter(provider=provider)
            .values("status")
            .annotate(count=Count("id"))
        )

        # Upcoming appointments (next 5)
        upcoming_appointments = list(
            Appointment.objects.filter(
                provider=provider, appointment_date__gte=now())
            .order_by("appointment_date")[:5]
            .values("id", "appointment_type", "appointment_date", "status",
                    patient_name=models.F("patient__first_name"))
        )

        # Recent visits (last 5)
        recent_visits = list(
            ExaminationRecord.objects.filter(provider=provider)
            .order_by("-visit_date")[:5]
            .values("id", "visit_date", mother_name=models.F("mother__first_name"))
        )

        # Recent mothers (last 5)
        recent_mothers = list(
            User.objects.filter(role="mother", created_by=provider)
            .order_by("-id")[:5]
            .values("id", "first_name", "last_name", "custom_id")
        )

        # Notifications count
        unread_notifications = Notification.objects.filter(
            user=provider, read=False
        ).count()

        data = {
            "metrics": metrics,
            "appointments_by_status": list(appointments_by_status),
            "upcoming_appointments": upcoming_appointments,
            "recent_visits": recent_visits,
            "recent_mothers": recent_mothers,
            "unread_notifications": unread_notifications,
        }

        serializer = ProviderDashboardSerializer(data)
        return Response(serializer.data)


class ProviderAppointmentsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        if user.role != "provider":
            return Response({"detail": "Forbidden"}, status=403)

        # Get all admin-created appointments
        admin_appts = Appointment.objects.all()

        # Get all provider-created appointments with a next appointment date
        exam_appts = ExaminationRecord.objects.filter(
            next_appointment__isnull=False)

        # Merge and serialize
        data = []
        for a in admin_appts:
            data.append(UnifiedAppointmentSerializer.from_admin(a))
        for e in exam_appts:
            data.append(UnifiedAppointmentSerializer.from_exam(e))

        # Sort by appointment_date (falling back to created_at)
        data = sorted(
            data, key=lambda x: x["appointment_date"] or x["created_at"], reverse=True
        )

        # Paginate
        paginator = PageNumberPagination()
        paginator.page_size = 10
        result_page = paginator.paginate_queryset(data, request)
        return paginator.get_paginated_response(result_page)

    def patch(self, request):
        """
        Update appointment status (works for both admin and provider-created appointments).
        """
        appt_id = request.data.get("id")
        source = request.data.get("source")
        new_status = request.data.get("status")

        if source == "admin":
            try:
                appt = Appointment.objects.get(id=appt_id)
                appt.status = new_status
                appt.save()
                return Response({"success": True, "msg": "Status updated"})
            except Appointment.DoesNotExist:
                return Response({"error": "Appointment not found"}, status=404)

        elif source == "provider":
            try:
                exam = ExaminationRecord.objects.get(id=appt_id)
                exam.status = new_status
                exam.save()
                return Response({"success": True, "msg": "Status updated"})
            except ExaminationRecord.DoesNotExist:
                return Response({"error": "Examination record not found"}, status=404)

        raise ValidationError({"error": "Invalid request"})
